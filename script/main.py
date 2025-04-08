
import os
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QLabel, QPushButton, QFileDialog, QTextEdit, QProgressBar,
                             QMessageBox)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QIcon
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from scipy.stats import linregress
from openpyxl.chart import ScatterChart, BarChart, Reference, Series
from openpyxl.chart.trendline import Trendline
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import __version__ as openpyxl_version
from openpyxl.utils import get_column_letter


# 美化样式
STYLE = """
QMainWindow {
    background-color: #2E2E2E;
}
QLabel {
    color: #FFFFFF;
    font-size: 14px;
}
QPushButton {
    background-color: #4CAF50;
    color: white;
    border: none;
    padding: 10px 24px;
    font-size: 14px;
    border-radius: 4px;
}
QPushButton:hover {
    background-color: #45a049;
}
QTextEdit {
    background-color: #FFFFFF;
    border: 1px solid #CCCCCC;
    border-radius: 4px;
    padding: 8px;
}
QProgressBar {
    border: 2px solid #444444;
    border-radius: 5px;
    text-align: center;
    background-color: #444444;
}
QProgressBar::chunk {
    background-color: #4CAF50;
    width: 10px;
}
"""

class AnalysisThread(QThread):
    progress_updated = pyqtSignal(int)
    log_message = pyqtSignal(str)
    analysis_complete = pyqtSignal(str)

    def __init__(self, input_path):
        super().__init__()
        self.input_path = input_path

    def run(self):
        try:
            self.log_message.emit("开始处理数据文件...")
            output_path = self.process_feeding_data()
            self.analysis_complete.emit(output_path)
        except Exception as e:
            self.log_message.emit(f"错误: {str(e)}")

    def process_feeding_data(self):
        # 自动生成输出路径
        input_dir = os.path.dirname(self.input_path)
        filename = os.path.basename(self.input_path)
        base_name, ext = os.path.splitext(filename)
        output_path = os.path.join(input_dir, f"{base_name}_RESULT{ext}")

        self.log_message.emit("正在加载Excel文件...")
        wb = load_workbook(self.input_path)  # 正确使用self.input_path
        sheet = wb["OD"]

    # ==================== 标准曲线处理 ====================
        self.log_message.emit("计算标准曲线...")
        std_data = [sheet.cell(row=row, column=3).value for row in range(25, 33)]
        concentrations = [0, 0.125, 0.25, 0.5, 0.75, 1, 1.5, 2]
    
    # 线性回归
        slope, intercept, r_value, _, _ = linregress(std_data, concentrations)
        r_squared = r_value**2

    # ==================== 计算C0 ====================
        self.log_message.emit("计算基准浓度C0...")
        c0_values = [sheet.cell(row=row, column=4).value for row in range(25, 33)]
        c0_conc = np.mean([slope*x + intercept for x in c0_values])

    # ==================== 处理样品数据 ====================
        self.log_message.emit("分析样品数据...")
        results = []
        V = 20  # 总液体体积

        for col_idx in range(5, 15):  # E-N列
            col_letter = chr(64 + col_idx)
            abs_values = [sheet.cell(row=row, column=col_idx).value for row in range(25, 33)]
        
            for row_idx, abs_val in enumerate(abs_values, 1):
                conc = slope * abs_val + intercept
                I = (1 - conc/c0_conc) * V
                results.append({
                    "Index": f"{col_letter}{row_idx}",
                    "Group": col_letter,
                    "OD": abs_val,
                    "C(μg/μl)": conc,
                    "Food intake (μl)": I
                })

        self.log_message.emit("生成结果文件...")
        result_df = pd.DataFrame(results)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 结果数据表
            result_df.to_excel(writer, index=False, sheet_name='Results')
            workbook = writer.book  # 必须先获取workbook对象
        # ==================== 统计汇总 ====================
            summary_df = result_df.groupby('Group')['Food intake (μl)'].agg(['mean', 'std', 'count'])
            summary_df.columns = ['Mean', 'Std Dev', 'Samples']
            summary_df.reset_index(inplace=True)

        # 写入统计结果
            summary_df.to_excel(writer, index=False, sheet_name='Summary')

        # ==================== 添加图表 ====================
        # 1. 添加柱状图
            summary_sheet = writer.sheets['Summary']
        
            bar_chart = BarChart()
            bar_chart.title = "Food Intake"
            bar_chart.y_axis.title = "Daily food intake per fly (μl)"
            bar_chart.x_axis.title = "Group"
        
            data = Reference(summary_sheet, 
                            min_col=2, 
                            min_row=1, 
                            max_col=3, 
                            max_row=len(summary_df)+1)
            categories = Reference(summary_sheet, 
                                min_col=1, 
                                min_row=2, 
                                max_row=len(summary_df)+1)
        
            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(categories)
        
            summary_sheet.add_chart(bar_chart, "F2")

        # 2. 添加标准曲线图
        
            std_sheet = workbook.create_sheet("Standard_Curve")
            from openpyxl.utils.dataframe import dataframe_to_rows
            std_df = pd.DataFrame({"OD": std_data, "Concentration": concentrations})
        
         # 写入标准曲线数据
            for r_idx, row in enumerate(dataframe_to_rows(std_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    std_sheet.cell(row=r_idx, column=c_idx, value=value)

        # 添加统计信息
            std_sheet['A11'] = "标准曲线公式:"
            std_sheet['B11'] = f"y = {slope:.4f}x + {intercept:.4f}"
            std_sheet['A12'] = "R²值:"
            std_sheet['B12'] = f"{r_squared:.4f}"
            std_sheet['A13'] = "C0值(μg/μl):"
            std_sheet['B13'] = f"{c0_conc:.2f}"

        # ==================== 添加标准曲线图表 ====================
            chart = ScatterChart()
            chart.title = "Standard Curve"
            chart.x_axis.title = 'OD Value'
            chart.y_axis.title = 'Concentration (μg/μl)'
            chart.width = 15
            chart.height = 8

        # 数据引用范围
            x_data = Reference(std_sheet, min_col=1, min_row=2, max_row=9)
            y_data = Reference(std_sheet, min_col=2, min_row=2, max_row=9)
            series = Series(y_data, x_data, title="Standard Data")
            chart.series.append(series)

        # 添加趋势线
            trendline = Trendline(trendlineType='linear')
            trendline.dispEq = True  
            trendline.dispRSqr = True  
        
        # 设置趋势线样式
            line_prop = LineProperties(solidFill="FF0000")
            trendline.graphicalProperties = GraphicalProperties(ln=line_prop)

        # 根据版本使用不同参数名
            if openpyxl_version.startswith('3.0'):
                trendline.graphicalProperties = GraphicalProperties(line=line_prop)
            else:
                trendline.graphicalProperties = GraphicalProperties(ln=line_prop)

            series.trendline = trendline

        # 将图表添加到工作表
            std_sheet.add_chart(chart, "D3")

        # ==================== 自动调整列宽 ====================
            for sheetname in writer.sheets:
                sheet = writer.sheets[sheetname]
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    sheet.column_dimensions[column_letter].width = adjusted_width

        return output_path

class BCAnalyzerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.setWindowIcon(QIcon("icon.ico"))  # 准备图标文件

    def initUI(self):
        self.setWindowTitle("吸光度数据分析系统")
        self.setGeometry(300, 300, 800, 600)
        self.setStyleSheet(STYLE)

        # 主控件
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)

        # 标题
        title = QLabel("吸光度数据分析系统")
        title.setFont(QFont("Arial", 18, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # 文件选择
        file_layout = QHBoxLayout()
        self.file_label = QLabel("选择输入文件:")
        self.file_path = QLabel("未选择文件")
        self.file_path.setStyleSheet("color: #FFFFFF;")
        browse_btn = QPushButton("浏览文件")
        browse_btn.clicked.connect(self.select_file)
        file_layout.addWidget(self.file_label)
        file_layout.addWidget(self.file_path)
        file_layout.addWidget(browse_btn)
        layout.addLayout(file_layout)

        # 进度条
        self.progress = QProgressBar()
        layout.addWidget(self.progress)

        # 日志显示
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        layout.addWidget(self.log)

        # 操作按钮
        self.analyze_btn = QPushButton("开始分析")
        self.analyze_btn.clicked.connect(self.start_analysis)
        layout.addWidget(self.analyze_btn)

    def select_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择数据文件", "", "Excel文件 (*.xlsx *.xls)"
        )
        if path:
            self.file_path.setText(path)
            self.log.append(f"已选择文件: {path}")

    def start_analysis(self):
        input_path = self.file_path.text()
        if not input_path or input_path == "未选择文件":
            QMessageBox.warning(self, "错误", "请先选择输入文件")
            return

        self.analyze_btn.setEnabled(False)
        self.log.append("\n开始处理数据...")

        self.thread = AnalysisThread(input_path)
        self.thread.log_message.connect(self.update_log)
        self.thread.progress_updated.connect(self.update_progress)
        self.thread.analysis_complete.connect(self.analysis_finished)
        self.thread.start()

    def update_log(self, message):
        self.log.append(message)

    def update_progress(self, value):
        self.progress.setValue(value)

    def analysis_finished(self, output_path):
        self.analyze_btn.setEnabled(True)
        self.log.append(f"\n分析完成！结果保存至: {output_path}")
        QMessageBox.information(self, "完成", "数据分析已完成！")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = BCAnalyzerApp()
    window.show()
    sys.exit(app.exec_())
