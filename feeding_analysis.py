import pandas as pd
import numpy as np
import os
from scipy.stats import linregress, f_oneway
from statsmodels.stats.multicomp import pairwise_tukeyhsd
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, BarChart, Reference, Series
from openpyxl.chart.trendline import Trendline
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl import __version__ as openpyxl_version

def process_feeding_data(input_path):
    # 自动生成输出路径
    input_dir = os.path.dirname(input_path)
    input_filename = os.path.basename(input_path)
    name, ext = os.path.splitext(input_filename)
    output_path = os.path.join(input_dir, f"{name}_RESULT{ext}")

    wb = load_workbook(input_path)
    sheet = wb["OD"]

    # ==================== 标准曲线处理 ====================
    std_data = [sheet.cell(row=row, column=3).value for row in range(25, 33)]
    concentrations = [0, 0.125, 0.25, 0.5, 0.75, 1, 1.5, 2]
    
    # 线性回归
    slope, intercept, r_value, _, _ = linregress(std_data, concentrations)
    r_squared = r_value**2

    # ==================== 计算C0 ====================
    c0_values = [sheet.cell(row=row, column=4).value for row in range(25, 33)]
    c0_conc = np.mean([slope*x + intercept for x in c0_values])

    # ==================== 处理样品数据 ====================
    results = []
    V = 20  # 总液体体积

    for col_idx in range(5, 15):  # E-N列
        col_letter = chr(64 + col_idx)
        abs_values = [sheet.cell(row=row, column=col_idx).value for row in range(25, 33)]
        
        for row_idx, abs_val in enumerate(abs_values, 1):
            conc = slope * abs_val + intercept
            I = (1 - conc/c0_conc) * V
            results.append({
                "Index": f"{col_letter}{row_idx}",
                "Group": col_letter,
                "OD": abs_val,
                "C(μg/μl)": conc,
                "Food intake (μl)": I
            })

    result_df = pd.DataFrame(results)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 写入结果数据
        result_df.to_excel(writer, index=False, sheet_name='Results')
        workbook = writer.book

        # ==================== 统计汇总 ====================
        summary_df = result_df.groupby('Group')['Food intake (μl)'].agg(['mean', 'std', 'count'])
        summary_df.columns = ['Mean', 'Std Dev', 'Samples']
        summary_df.reset_index(inplace=True)

        # 写入统计结果
        summary_df.to_excel(writer, index=False, sheet_name='Summary')

        # ==================== 添加图表 ====================
        # 1. 添加柱状图
        summary_sheet = writer.sheets['Summary']
        
        bar_chart = BarChart()
        bar_chart.title = "Food Intake"
        bar_chart.y_axis.title = "Daily food intake per fly (μl)"
        bar_chart.x_axis.title = "Group"
        
        data = Reference(summary_sheet, 
                        min_col=2, 
                        min_row=1, 
                        max_col=3, 
                        max_row=len(summary_df)+1)
        categories = Reference(summary_sheet, 
                              min_col=1, 
                              min_row=2, 
                              max_row=len(summary_df)+1)
        
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(categories)
        
        summary_sheet.add_chart(bar_chart, "F2")

        # 2. 添加标准曲线图
        
        std_sheet = workbook.create_sheet("Standard_Curve")
        std_df = pd.DataFrame({"OD": std_data, "Concentration": concentrations})
        
         # 写入标准曲线数据
        for row in dataframe_to_rows(std_df, index=False, header=True):
            std_sheet.append(row)

        # 添加统计信息
        std_sheet['A11'] = "标准曲线公式:"
        std_sheet['B11'] = f"y = {slope:.4f}x + {intercept:.4f}"
        std_sheet['A12'] = "R²值:"
        std_sheet['B12'] = f"{r_squared:.4f}"
        std_sheet['A13'] = "C0值(μg/μl):"
        std_sheet['B13'] = f"{c0_conc:.2f}"

        # ==================== 添加标准曲线图表 ====================
        chart = ScatterChart()
        chart.title = "Standard Curve"
        chart.x_axis.title = 'OD Value'
        chart.y_axis.title = 'Concentration (μg/μl)'
        chart.width = 15
        chart.height = 8

        # 数据引用范围
        x_data = Reference(std_sheet, min_col=1, min_row=2, max_row=9)
        y_data = Reference(std_sheet, min_col=2, min_row=2, max_row=9)
        series = Series(y_data, x_data, title="Standard Data")
        chart.series.append(series)

        # 添加趋势线
        trendline = Trendline(trendlineType='linear')
        trendline.dispEq = True  
        trendline.dispRSqr = True  
        
        # 设置趋势线样式
        line_prop = LineProperties(solidFill=ColorChoice(prstClr='red'))

        # 根据版本使用不同参数名
        if openpyxl_version.startswith('3.0'):
            trendline.graphicalProperties = GraphicalProperties(line=line_prop)
        else:
            trendline.graphicalProperties = GraphicalProperties(ln=line_prop)

        series.trendline = trendline

        # 将图表添加到工作表
        std_sheet.add_chart(chart, "D3")

        # ==================== 自动调整列宽 ====================
        for sheetname in writer.sheets:
            sheet = writer.sheets[sheetname]
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width

    return slope, intercept, r_squared, c0_conc, output_path

# 使用示例
if __name__ == "__main__":
    input_path = r"E:\py_zhz\BCA\BCA_DATA_PROCESS\20250218.OD.xlsx"
    slope, intercept, r_squared, c0_conc, output_path = process_feeding_data(input_path)
    
    print(f"处理完成！结果保存至：{output_path}")
    print(f"标准曲线方程: y = {slope:.4f}x + {intercept:.4f}")
    print(f"R²值: {r_squared:.4f}")
    print(f"C0基准浓度: {c0_conc:.2f} μg/μl")